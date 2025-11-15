"""Web Scraper for Excel Automation
This script scrapes data from websites and automatically exports to Excel.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log'),
        logging.StreamHandler()
    ]
)

class WebScraper:
    """Class to handle web scraping and Excel export operations"""
    
    def __init__(self, url, output_file='scraped_data.xlsx'):
        """
        Initialize the scraper
        Args:
            url (str): Target URL to scrape
            output_file (str): Output Excel filename
        """
        self.url = url
        self.output_file = output_file
        self.data = []
        
    def fetch_page(self):
        """
        Fetch webpage content
        Returns:
            BeautifulSoup object or None if error
        """
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(self.url, headers=headers, timeout=10)
            response.raise_for_status()
            logging.info(f"Successfully fetched: {self.url}")
            return BeautifulSoup(response.content, 'html.parser')
        except requests.RequestException as e:
            logging.error(f"Error fetching {self.url}: {e}")
            return None
    
    def parse_data(self, soup, selector='.data-item'):
        """
        Parse data from HTML using CSS selector
        Args:
            soup (BeautifulSoup): Parsed HTML
            selector (str): CSS selector for target elements
        Returns:
            list: Extracted data
        """
        try:
            items = soup.select(selector)
            for item in items:
                text = item.get_text(strip=True)
                if text:
                    self.data.append([text])
            logging.info(f"Extracted {len(self.data)} items")
            return self.data
        except Exception as e:
            logging.error(f"Error parsing data: {e}")
            return []
    
    def export_to_excel(self):
        """
        Export scraped data to Excel with formatting
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Scraped Data'
            
            # Add header
            ws['A1'] = 'Scraped Content'
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add data
            for idx, row in enumerate(self.data, start=2):
                ws.append(row)
            
            # Add metadata
            ws['C1'] = 'Scraped on:'
            ws['D1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['C2'] = 'Source URL:'
            ws['D2'] = self.url
            
            # Adjust column width
            ws.column_dimensions['A'].width = 50
            
            wb.save(self.output_file)
            logging.info(f"Data exported to {self.output_file}")
            return True
        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}")
            return False
    
    def run(self, selector='.data-item'):
        """
        Execute complete scraping workflow
        Args:
            selector (str): CSS selector for target data
        """
        logging.info("Starting web scraping process...")
        soup = self.fetch_page()
        if soup:
            self.parse_data(soup, selector)
            if self.data:
                self.export_to_excel()
                logging.info("Scraping completed successfully!")
            else:
                logging.warning("No data extracted. Check the selector.")
        else:
            logging.error("Failed to fetch webpage")


if __name__ == "__main__":
    # Example usage
    target_url = 'https://example.com/data-page'
    css_selector = '.target-class'  # Update with actual selector
    
    scraper = WebScraper(target_url, output_file='output.xlsx')
    scraper.run(selector=css_selector)
    
    print(f"\nScraping complete! Check output.xlsx and scraper.log for details.")
